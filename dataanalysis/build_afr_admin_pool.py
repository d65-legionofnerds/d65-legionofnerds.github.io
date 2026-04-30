"""
Parse all D65 AFR files to extract the central-office and school-administration
compensation pool by function code over time.

This complements the PA-disclosure analysis by providing a salary-threshold-free
view of admin compensation. The AFR captures every employee in admin functions
regardless of salary, so it includes sub-$75K admin support staff that the
PA reports omit.

Functions counted:
  2300 series — General Administration (Supt., Asst Supts, central admin, board, curriculum dirs/coords)
  2400 series — School Administration (principals, asst principals, school office support)
  2500 series — Business Services (CFO, fiscal, payroll; excludes O&M)
  2600 series — Central Support (planning, IT, staff services, data processing)

Admin compensation = Salaries (object 100) + Employee Benefits (object 200)
                     + MR/SS pension benefits in the same function code
                     (these appear in the SUPPORT SERVICES (MR/SS) section).

Output:
  data/afr_admin_pool.csv  — long-format with year, function, sal, ben, mrss
  data/afr_admin_pool_summary.csv  — wide with year-by-year totals

Reads:
  data/afr/*.xlsx and *.xls

Run:
  PYTHONIOENCODING=utf-8 python build_afr_admin_pool.py
"""
import os
import re
import warnings
import openpyxl
import xlrd
import pandas as pd

warnings.filterwarnings("ignore")

ROOT = os.path.dirname(os.path.abspath(__file__))
AFR_DIR = os.path.join(ROOT, "data", "afr")
OUT_DIR = os.path.join(ROOT, "data")

# Function codes we treat as central-office + school administration.
ADMIN_FUNCTIONS = {
    "2300": "General Administration (total)",
    "2310": "Board of Education",
    "2320": "Executive Administration",
    "2330": "Special Area Administration",
    "2400": "School Administration (total)",
    "2410": "Office of the Principal",
    "2490": "Other School Administration",
    "2500": "Business Services (total)",
    "2510": "Direction of Business Support Services",
    "2520": "Fiscal Services",
    "2540": "Operation & Maintenance",
    "2550": "Pupil Transportation",
    "2560": "Food Services",
    "2570": "Internal Services",
    "2600": "Central Support Services (total)",
    "2610": "Direction of Central Support Services",
    "2620": "Planning, Research, Development, Evaluation",
    "2630": "Information Services",
    "2640": "Staff Services",
    "2660": "Data Processing",
}

ADMIN_FUNCTION_GROUPS = {
    "2300": "General Administration",
    "2400": "School Administration",
    "2500": "Business Services",
    "2600": "Central Support Services",
}


def fy_from_filename(filename):
    """Extract two-digit FY from filename like AFR12 / AFR25."""
    m = re.search(r"AFR\s*(\d{2})", filename, re.IGNORECASE)
    if m:
        return 2000 + int(m.group(1))
    m = re.search(r"Budget\s*(\d{2})", filename, re.IGNORECASE)
    if m:
        return 2000 + int(m.group(1))
    return None


def parse_xlsx(fp):
    wb = openpyxl.load_workbook(fp, data_only=True)
    expsheet = None
    for n in wb.sheetnames:
        if "Expenditures" in n:
            expsheet = wb[n]
            break
    if expsheet is None:
        return None
    rows = []
    for r in range(1, expsheet.max_row + 1):
        rowvals = [expsheet.cell(row=r, column=c).value for c in range(1, 12)]
        rows.append(rowvals)
    return rows


def parse_xls(fp):
    wb = xlrd.open_workbook(fp)
    expsheet = None
    for n in wb.sheet_names():
        if "Expenditures" in n:
            expsheet = wb.sheet_by_name(n)
            break
    if expsheet is None:
        return None
    rows = []
    for r in range(expsheet.nrows):
        rowvals = []
        for c in range(min(11, expsheet.ncols)):
            v = expsheet.cell_value(r, c)
            if v == "":
                v = None
            rowvals.append(v)
        rows.append(rowvals)
    return rows


def normalize_func(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    # Strip trailing .0
    if s.endswith(".0"):
        s = s[:-2]
    return s


def num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        v2 = v.replace(",", "").replace("$", "").strip()
        if v2 == "" or v2 == "-":
            return 0.0
        try:
            return float(v2)
        except ValueError:
            return 0.0
    return 0.0


def parse_afr(filepath, year):
    """Walk the Expenditures sheet and capture function-level admin spending.

    Returns a list of dicts: {year, fund_section, func, salaries, benefits, ...}
    The MR/SS section reports benefits-only with the same function codes.
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        rows = parse_xlsx(filepath)
    elif ext == ".xls":
        rows = parse_xls(filepath)
    else:
        return []

    if rows is None:
        return []

    records = []
    current_section = ""
    for row in rows:
        # row layout: [desc, func#, salaries, benefits, purch, supplies, capital, other, noncap, term, total]
        desc = row[0]
        if isinstance(desc, str):
            d = desc.strip()
            if "EDUCATIONAL FUND" in d or "OPERATIONS & MAINTENANCE" in d or "MUNICIPAL RETIREMENT" in d \
                    or "TRANSPORTATION FUND" in d or "TORT" in d or "CAPITAL PROJECTS" in d \
                    or "FIRE PREVENTION" in d:
                current_section = d
            elif "SUPPORT SERVICES (" in d:
                current_section = d.strip()
            elif "INSTRUCTION (" in d:
                current_section = d.strip()

        func = normalize_func(row[1])
        if func not in ADMIN_FUNCTIONS:
            continue

        sal = num(row[2])
        ben = num(row[3])
        purch = num(row[4]) if len(row) > 4 else 0.0
        total = num(row[10]) if len(row) > 10 else 0.0

        records.append({
            "year": year,
            "filename": os.path.basename(filepath),
            "section": current_section,
            "func": func,
            "func_label": ADMIN_FUNCTIONS[func],
            "salaries": sal,
            "benefits": ben,
            "purchased_services": purch,
            "row_total": total,
        })
    return records


def main():
    files = sorted(os.listdir(AFR_DIR))
    all_records = []
    for f in files:
        fy = fy_from_filename(f)
        if fy is None:
            continue
        # Skip the duplicate AFR25 (we have two: AFR25_d65.xlsx and the longer name).
        if f == "AFR25_d65.xlsx":
            continue
        fp = os.path.join(AFR_DIR, f)
        recs = parse_afr(fp, fy)
        print(f"{f:80s}  FY{fy}  {len(recs):3d} admin function rows")
        all_records.extend(recs)

    df = pd.DataFrame(all_records)
    df.to_csv(os.path.join(OUT_DIR, "afr_admin_pool.csv"), index=False)
    print(f"\nWrote afr_admin_pool.csv with {len(df)} rows")

    # Build a summary: for each year, total comp (salaries+benefits) for each
    # admin function group (2300, 2400, 2500, 2600). MR/SS pension contributions
    # appear in the same function codes within the SUPPORT SERVICES (MR/SS)
    # section, but only as "benefits". We sum those into mrss.
    # Heuristic: rows where salaries==0 and section contains "MR/SS" are pension benefits.
    df["is_mrss"] = df["section"].str.contains("MR/SS", na=False)
    df["function_group"] = df["func"].str[0:2] + "00"

    # For top-level function totals (2300, 2400, 2500, 2600), only use rows where
    # func == group code (i.e., the row that says "Total Support Services - General Administration").
    # This avoids double-counting subfunction rows.
    top_level = df[df["func"].isin(["2300", "2400", "2500", "2600"])].copy()

    # Education-fund admin (excluding MR/SS section): salaries+benefits
    ed_admin = (top_level[~top_level["is_mrss"] & top_level["section"].str.contains("\\(ED\\)", na=False)]
                .groupby(["year", "func"], as_index=False)
                .agg(salaries=("salaries", "sum"),
                     benefits=("benefits", "sum")))
    # MR/SS pension benefits in admin functions
    mrss_admin = (top_level[top_level["is_mrss"]]
                  .groupby(["year", "func"], as_index=False)
                  .agg(mrss=("benefits", "sum")))

    summary = ed_admin.merge(mrss_admin, on=["year", "func"], how="left").fillna(0.0)
    summary["total_comp"] = summary["salaries"] + summary["benefits"] + summary["mrss"]
    summary["function_label"] = summary["func"].map({
        "2300": "General Administration",
        "2400": "School Administration",
        "2500": "Business Services",
        "2600": "Central Support Services",
    })

    # Pivot for easy reading
    pivot = summary.pivot_table(index="year", columns="func", values="total_comp", aggfunc="sum").reset_index()
    pivot.columns.name = None
    pivot["pool_total"] = pivot[["2300", "2400", "2500", "2600"]].sum(axis=1, numeric_only=True)
    pivot["pool_excl_principals"] = pivot[["2300", "2500", "2600"]].sum(axis=1, numeric_only=True)

    # Apply CPI to compute real (2026) dollars
    cpi = pd.read_csv(os.path.join(OUT_DIR, "cpi_history.csv"))
    cpi_2026 = cpi.loc[cpi["year"] == 2026, "cpi_u"].iloc[0]
    cpi["adj_to_2026"] = cpi_2026 / cpi["cpi_u"]
    pivot = pivot.merge(cpi[["year", "adj_to_2026"]], on="year", how="left")
    for col in ["2300", "2400", "2500", "2600", "pool_total", "pool_excl_principals"]:
        pivot[f"{col}_real"] = pivot[col] * pivot["adj_to_2026"]

    pivot.to_csv(os.path.join(OUT_DIR, "afr_admin_pool_summary.csv"), index=False)
    print(f"Wrote afr_admin_pool_summary.csv with {len(pivot)} years")

    print("\n=== AFR Admin Pool Summary (in 2026 $) ===")
    cols = ["year", "2300_real", "2400_real", "2500_real", "2600_real", "pool_total_real", "pool_excl_principals_real"]
    out = pivot[cols].copy()
    for c in cols[1:]:
        out[c] = out[c].apply(lambda v: f"${v/1e6:>5.2f}M" if pd.notna(v) else "")
    print(out.to_string(index=False))


if __name__ == "__main__":
    main()
