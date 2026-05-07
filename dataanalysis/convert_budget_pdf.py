"""
Convert the FY2026 D65 ISBE Budget PDF to an xlsx file matching the AFR format
so that build_budget_explorer.py can parse it.

Input:  C:\\Users\\jkarlin\\Downloads\\D65ISBEBudget09292025.pdf
Output: dataanalysis/data/afr/05-016-0650-04_AFR26 Evanston CCSD 65.xlsx
"""

import os
import re

import openpyxl
import pdfplumber

PDF_PATH = r"C:\Users\jkarlin\Downloads\D65ISBEBudget09292025.pdf"
ROOT = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(ROOT, "data", "afr",
                        "05-016-0650-04_AFR26 Evanston CCSD 65.xlsx")


def clean_num(val):
    if val is None:
        return 0
    s = str(val).strip().replace(",", "").replace("$", "")
    s = s.replace("(", "-").replace(")", "")
    if s == "" or s == "-":
        return 0
    try:
        return float(s)
    except ValueError:
        return 0


def parse_func(val):
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace(" ", "")
    if s in ("2361,2365", "23612365", "2361\n2365"):
        return 2361
    s = s.split("\n")[0].strip()
    m = re.match(r"^(\d{3,4})$", s)
    if m:
        return int(m.group(1))
    return None


def is_expenditure_page(page):
    text = page.extract_text() or ""
    return "Estimated Disbursements/Expenditures" in text


def extract_all_tables(pdf):
    """Extract tables from all expenditure pages, returning clean rows.

    pdfplumber's line-based extraction gives us 12 columns per row:
      [row_num/empty, description, funct#, salaries, benefits, purch_services,
       supplies, capital, other, noncap, termination, total]
    """
    all_rows = []

    for pi, page in enumerate(pdf.pages):
        if not is_expenditure_page(page):
            continue

        tables = page.extract_tables({
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "snap_tolerance": 5,
        })

        for table in tables:
            for ri, row in enumerate(table):
                if row is None or len(row) < 4:
                    continue
                if ri == 0:
                    first_cell = str(row[0] or "")
                    if len(first_cell) > 200:
                        continue

                desc = str(row[1] or "").strip() if len(row) > 1 else ""
                if not desc:
                    continue

                func = parse_func(row[2]) if len(row) > 2 else None

                nums = []
                for i in range(3, min(len(row), 12)):
                    nums.append(clean_num(row[i]))
                while len(nums) < 9:
                    nums.append(0)

                out_row = [desc, func] + nums[:9]
                all_rows.append(out_row)

    return all_rows


def main():
    print(f"Reading {PDF_PATH}")
    pdf = pdfplumber.open(PDF_PATH)

    rows = extract_all_tables(pdf)
    pdf.close()
    print(f"Extracted {len(rows)} data rows from expenditure pages")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenditures 16-24"

    for row in rows:
        ws.append(row)

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")

    verify()


def verify():
    print("\n=== Verification ===")
    wb = openpyxl.load_workbook(OUT_PATH, data_only=True)
    ws = wb.active
    print(f"Sheet: {ws.title}, Rows: {ws.max_row}")

    current_section = ""
    ed_fund_codes = []

    for r in range(1, ws.max_row + 1):
        desc = str(ws.cell(row=r, column=1).value or "").upper()
        func = ws.cell(row=r, column=2).value
        total = ws.cell(row=r, column=11).value

        if "EDUCATIONAL FUND" in desc:
            current_section = "ED"
        elif "OPERATIONS AND MAINTENANCE" in desc or "O&M" in desc.replace(" ", ""):
            current_section = "OM"
        elif "MUNICIPAL RETIREMENT" in desc or "MR/SS" in desc:
            current_section = "MRSS"
        elif "TRANSPORTATION FUND" in desc or "(TR)" in desc:
            current_section = "TR"
        elif "DEBT SERVICE FUND" in desc or "(DS)" in desc:
            current_section = "DS"
        elif "TORT" in desc and "FUND" in desc:
            current_section = "TORT"

        if current_section == "ED" and func is not None:
            ed_fund_codes.append((func, desc.strip()[:50], total))

    print(f"\nEd Fund function codes found: {len(ed_fund_codes)}")
    key_codes = [1100, 1200, 1800, 2110, 2310, 2410, 2660, 3000, 4120, 4220]
    for fc, desc, total in ed_fund_codes:
        if fc in key_codes:
            print(f"  {fc:>6}  total={total:>12}  {desc}")

    ed_total = sum(t for _, _, t in ed_fund_codes
                   if t is not None and isinstance(t, (int, float)))
    print(f"\n  Sum of all Ed Fund rows (including rollups): ${ed_total:,.0f}")


if __name__ == "__main__":
    main()
