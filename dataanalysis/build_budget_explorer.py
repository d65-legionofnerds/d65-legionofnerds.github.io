"""
Parse D65 AND comparable peer-district AFR files to extract Educational Fund
expenditures by ISBE function code, then generate a standalone interactive HTML
budget explorer that lets the user compare D65 against peer districts.

Peers are the same comparable districts used in the enrollment analysis
(enrollment-data.md / calculations.ipynb target_districts). Their AFRs are
downloaded by download_peer_afrs.py into data/afr/peers/{slug}/.

For each district and year we capture:
  - every Educational Fund function-code expenditure total
  - the district's total Ed Fund spend (denominator for "% of budget")
  - the official 9-month ADA from the PCTC-OEPP sheet (denominator for per-pupil)

On-behalf / pension payments are excluded.

Output:
  assets/budget_explorer.html — standalone interactive Plotly app

Run:
  python download_peer_afrs.py     # one-time, to fetch peer AFRs
  python build_budget_explorer.py
"""

import json
import os
import re
import warnings

import openpyxl
import pandas as pd
import xlrd

warnings.filterwarnings("ignore")

ROOT = os.path.dirname(os.path.abspath(__file__))
AFR_DIR = os.path.join(ROOT, "data", "afr")
PEER_DIR = os.path.join(AFR_DIR, "peers")
ASSETS_DIR = os.path.join(ROOT, "assets")

# D65 first, then peers. slug -> display name. Peer slugs are the
# districtvitals subdomains / folder names under data/afr/peers/.
D65_KEY = "d65"
D65_NAME = "Evanston CCSD 65"

PEER_NAMES = {
    "eastmained63": "East Maine SD 63",
    "winnetkad36": "Winnetka SD 36",
    "sd28": "Northbrook SD 28",
    "sd35": "Glencoe SD 35",
    "ccsd62": "CCSD 62 (Des Plaines)",
    "ccsd64": "Park Ridge CCSD 64",
    "lincolnwoodd74": "Lincolnwood SD 74",
    "arlingtonheightsd25": "Arlington Heights SD 25",
    "skokied68": "Skokie SD 68",
    "skokied69": "Skokie SD 69",
    "sd735": "Skokie SD 73-5",
    "esd97": "Oak Park ESD 97",
    "northbrookglenviewd30": "Northbrook/Glenview SD 30",
    "glenviewd34": "Glenview CCSD 34",
    "sd39": "Wilmette SD 39",
    "ccsd21": "Wheeling CCSD 21",
    "palatined15": "Palatine CCSD 15",
    "sd112": "North Shore SD 112",
}

FUNCTION_HIERARCHY = {
    "1100": ("Instruction", "Regular Programs", "Regular Programs K-12"),
    "1115": ("Instruction", "Regular Programs", "Charter Schools"),
    "1125": ("Instruction", "Pre-K Programs", "Pre-K Programs"),
    "1200": ("Instruction", "Special Education", "Special Ed K-12"),
    "1225": ("Instruction", "Special Education", "Special Ed Pre-K"),
    "1250": ("Instruction", "Remedial/Supplemental", "Remedial & Supplemental K-12"),
    "1275": ("Instruction", "Remedial/Supplemental", "Remedial & Supplemental Pre-K"),
    "1300": ("Instruction", "Other Instruction", "Adult/Continuing Ed"),
    "1400": ("Instruction", "Other Instruction", "CTE Programs"),
    "1500": ("Instruction", "Co-Curricular", "Interscholastic Programs"),
    "1600": ("Instruction", "Summer & Extended", "Summer School"),
    "1650": ("Instruction", "Gifted", "Gifted Programs"),
    "1700": ("Instruction", "Other Instruction", "Driver's Education"),
    "1800": ("Instruction", "Bilingual", "Bilingual Programs"),
    "1900": ("Instruction", "Other Instruction", "Truant Alt & Optional Programs"),
    "1910": ("Instruction", "Private Tuition", "Private Tuition - Regular Programs"),
    "1911": ("Instruction", "Private Tuition", "Private Tuition - Special Ed"),
    "1912": ("Instruction", "Private Tuition", "Private Tuition - CTE"),
    "1913": ("Instruction", "Private Tuition", "Private Tuition - Other"),
    "1914": ("Instruction", "Private Tuition", "Private Tuition - Interscholastic"),
    "1915": ("Instruction", "Private Tuition", "Private Tuition - Summer School"),
    "1916": ("Instruction", "Private Tuition", "Private Tuition - Gifted"),
    "1917": ("Instruction", "Private Tuition", "Private Tuition - Driver's Ed"),
    "1918": ("Instruction", "Private Tuition", "Private Tuition - Bilingual"),
    "1919": ("Instruction", "Private Tuition", "Private Tuition - Truant Alt"),
    "1920": ("Instruction", "Private Tuition", "Private Tuition - Pre-K"),
    "1921": ("Instruction", "Private Tuition", "Private Tuition - Remedial"),
    "1922": ("Instruction", "Private Tuition", "Private Tuition - Spec Ed Pre-K"),
    "1999": ("Instruction", "Other Instruction", "Student Activity Fund"),
    "2110": ("Student Support", "Student Support", "Attendance & Social Work"),
    "2120": ("Student Support", "Student Support", "Guidance Services"),
    "2130": ("Student Support", "Student Support", "Health Services"),
    "2140": ("Student Support", "Student Support", "Psychological Services"),
    "2150": ("Student Support", "Student Support", "Speech Pathology & Audiology"),
    "2190": ("Student Support", "Student Support", "Other Student Support"),
    "2210": ("Instructional Staff Support", "Instructional Staff Support", "Improvement of Instruction"),
    "2220": ("Instructional Staff Support", "Instructional Staff Support", "Educational Media"),
    "2230": ("Instructional Staff Support", "Instructional Staff Support", "Assessment & Testing"),
    "2310": ("General Administration", "General Administration", "Board of Education"),
    "2320": ("General Administration", "General Administration", "Executive Administration"),
    "2330": ("General Administration", "General Administration", "Special Area Administration"),
    "2361": ("General Administration", "General Administration", "Tort Immunity"),
    "2365": ("General Administration", "General Administration", "Tort/Worker's Comp"),
    "2410": ("School Administration", "School Administration", "Office of the Principal"),
    "2490": ("School Administration", "School Administration", "Other School Admin"),
    "2510": ("Business Services", "Business Services", "Direction of Business Support"),
    "2520": ("Business Services", "Business Services", "Fiscal Services"),
    "2540": ("Business Services", "Business Services", "Operation & Maintenance of Plant"),
    "2550": ("Business Services", "Business Services", "Pupil Transportation"),
    "2560": ("Business Services", "Business Services", "Food Services"),
    "2570": ("Business Services", "Business Services", "Internal Services"),
    "2610": ("Central Support", "Central Support", "Direction of Central Support"),
    "2620": ("Central Support", "Central Support", "Planning, Research & Development"),
    "2630": ("Central Support", "Central Support", "Information Services"),
    "2640": ("Central Support", "Central Support", "Staff Services"),
    "2660": ("Central Support", "Central Support", "Data Processing / IT"),
    "2900": ("Other Support", "Other Support", "Other Support Services"),
    "3000": ("Community Services", "Community Services", "Community Services"),
    "4110": ("Payments to Other Districts", "In-State Payments", "Payments - Regular Programs"),
    "4120": ("Payments to Other Districts", "In-State Payments", "Payments - Special Ed"),
    "4130": ("Payments to Other Districts", "In-State Payments", "Payments - Adult/Continuing Ed"),
    "4140": ("Payments to Other Districts", "In-State Payments", "Payments - CTE"),
    "4170": ("Payments to Other Districts", "In-State Payments", "Payments - Community College"),
    "4190": ("Payments to Other Districts", "In-State Payments", "Other In-State Payments"),
    "4210": ("Payments to Other Districts", "Tuition Payments", "Tuition - Regular Programs"),
    "4220": ("Payments to Other Districts", "Tuition Payments", "Tuition - Special Ed"),
    "4230": ("Payments to Other Districts", "Tuition Payments", "Tuition - Adult/Continuing Ed"),
    "4240": ("Payments to Other Districts", "Tuition Payments", "Tuition - CTE"),
    "4270": ("Payments to Other Districts", "Tuition Payments", "Tuition - Community College"),
    "4280": ("Payments to Other Districts", "Tuition Payments", "Tuition - Other Programs"),
    "4290": ("Payments to Other Districts", "Tuition Payments", "Other Tuition Payments"),
    "4310": ("Payments to Other Districts", "Transfer Payments", "Transfers - Regular Programs"),
    "4320": ("Payments to Other Districts", "Transfer Payments", "Transfers - Special Ed"),
    "4330": ("Payments to Other Districts", "Transfer Payments", "Transfers - Adult/Continuing Ed"),
    "4340": ("Payments to Other Districts", "Transfer Payments", "Transfers - CTE"),
    "4370": ("Payments to Other Districts", "Transfer Payments", "Transfers - Community College"),
    "4380": ("Payments to Other Districts", "Transfer Payments", "Transfers - Other Programs"),
    "4390": ("Payments to Other Districts", "Transfer Payments", "Other Transfer Payments"),
    "4400": ("Payments to Other Districts", "Out-of-State Payments", "Payments to Out-of-State"),
    "5110": ("Debt Service", "Debt Service", "Interest - Tax Anticipation Warrants"),
    "5120": ("Debt Service", "Debt Service", "Interest - Tax Anticipation Notes"),
    "5130": ("Debt Service", "Debt Service", "Interest - CPPRT Anticipation Notes"),
    "5140": ("Debt Service", "Debt Service", "Interest - Other Short-Term"),
    "5150": ("Debt Service", "Debt Service", "Other Interest on Short-Term Debt"),
    "5200": ("Debt Service", "Debt Service", "Interest on Long-Term Debt"),
    "5300": ("Debt Service", "Debt Service", "Principal on Long-Term Debt"),
    "5400": ("Debt Service", "Debt Service", "Debt Service - Other"),
    "6000": ("Contingencies", "Contingencies", "Provisions for Contingencies"),
}

ROLLUP_CODES = {
    "1000", "2000", "2100", "2200", "2300", "2400", "2500", "2600",
    "3000", "4000", "4100", "4200", "4300", "5000", "5100",
}

KNOWN_LEAF_CODES = set(FUNCTION_HIERARCHY.keys())

EXCLUDE_SECTION_KEYWORDS = [
    "MR/SS", "O&M", "TORT FUND", "TRANSPORTATION FUND",
    "CAPITAL PROJECTS", "FIRE PREVENTION", "MUNICIPAL RETIREMENT",
    "OPERATIONS & MAINTENANCE",
]


# ---------------------------------------------------------------------------
# Excel parsing — robust to mislabeled extensions (some ISBE files have a .xls
# extension but xlsx content, or vice versa). Detect by magic bytes.
# ---------------------------------------------------------------------------

def detect_format(fp):
    with open(fp, "rb") as fh:
        head = fh.read(8)
    if head[:4] == b"PK\x03\x04":
        return "xlsx"
    if head[:4] == b"\xd0\xcf\x11\xe0":
        return "xls"
    # Fall back to extension.
    return "xlsx" if os.path.splitext(fp)[1].lower() in (".xlsx", ".xlsm") else "xls"


def fy_from_filename(filename):
    m = re.search(r"AFR\s*(\d{2})", filename, re.IGNORECASE)
    if m:
        return 2000 + int(m.group(1))
    m = re.search(r"Budget\s*(\d{2})", filename, re.IGNORECASE)
    if m:
        return 2000 + int(m.group(1))
    return None


def _sheet_rows_xlsx(fp, name_pred, ncols=12):
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    sheet = None
    for n in wb.sheetnames:
        if name_pred(n):
            sheet = wb[n]
            break
    if sheet is None:
        wb.close()
        return None
    rows = []
    for r in sheet.iter_rows(min_col=1, max_col=ncols, values_only=True):
        rows.append(list(r))
    wb.close()
    return rows


def _sheet_rows_xls(fp, name_pred, ncols=12):
    wb = xlrd.open_workbook(fp)
    sheet = None
    for n in wb.sheet_names():
        if name_pred(n):
            sheet = wb.sheet_by_name(n)
            break
    if sheet is None:
        return None
    rows = []
    for r in range(sheet.nrows):
        rowvals = []
        for c in range(min(ncols, sheet.ncols)):
            v = sheet.cell_value(r, c)
            if v == "":
                v = None
            rowvals.append(v)
        rows.append(rowvals)
    return rows


def read_sheet(fp, name_pred, ncols=12):
    """Return list-of-rows for the first sheet whose name matches name_pred."""
    fmt = detect_format(fp)
    try:
        if fmt == "xlsx":
            return _sheet_rows_xlsx(fp, name_pred, ncols)
        return _sheet_rows_xls(fp, name_pred, ncols)
    except Exception:
        # Try the other engine if the magic-byte guess was wrong.
        try:
            if fmt == "xlsx":
                return _sheet_rows_xls(fp, name_pred, ncols)
            return _sheet_rows_xlsx(fp, name_pred, ncols)
        except Exception:
            return None


def normalize_func(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        v2 = v.replace(",", "").replace("$", "").strip()
        if v2 == "" or v2 == "-":
            return 0.0
        try:
            return float(v2)
        except ValueError:
            return 0.0
    return 0.0


# ---------------------------------------------------------------------------
# AFR parsing — all Ed Fund function codes + 9-month ADA
# ---------------------------------------------------------------------------

def is_ed_section(section):
    if "(ED)" in section:
        return True
    if "EDUCATIONAL FUND" in section:
        return True
    return False


def is_excluded_section(section):
    s = section.upper()
    for kw in EXCLUDE_SECTION_KEYWORDS:
        if kw in s:
            return True
    return False


FUND_MARKERS = re.compile(
    r"\((ED|O&M|MR/SS|TORT|TR|DS|CP|FP&S|TF)\)", re.IGNORECASE
)

ADA_RE = re.compile(r"\bADA\b", re.IGNORECASE)
ADA_9MO_RE = re.compile(r"9\s*-?\s*MO|9\s*MONTH|NINE\s*MONTH", re.IGNORECASE)


def parse_afr_all(filepath, year):
    rows = read_sheet(filepath, lambda n: "Expenditures" in n)
    if rows is None:
        return [], {}

    records = []
    rollups = {}
    current_section = ""

    for row in rows:
        desc = row[0]
        if isinstance(desc, str):
            d = desc.strip().upper()
            if "EDUCATIONAL FUND" in d or "OPERATIONS & MAINTENANCE FUND" in d:
                current_section = d
            elif FUND_MARKERS.search(d):
                current_section = d

        if is_excluded_section(current_section) or not is_ed_section(current_section):
            continue

        if isinstance(desc, str) and ("ON BEHALF" in desc.upper() or "ON-BEHALF" in desc.upper()):
            continue

        func = normalize_func(row[1])
        if func is None:
            continue

        total = num(row[10]) if len(row) > 10 else 0.0
        salaries = num(row[2])
        benefits = num(row[3])

        if func in ROLLUP_CODES and func != "3000":
            rollups[func] = total
            continue

        if func in KNOWN_LEAF_CODES or (func.isdigit() and len(func) == 4):
            records.append({
                "func": func,
                "total": total,
                "salaries": salaries,
                "benefits": benefits,
            })

    return records, rollups


def extract_ada(filepath):
    """Pull the official 9-month ADA from the PCTC-OEPP sheet (per-pupil
    denominator). Returns a float or None."""
    rows = read_sheet(
        filepath,
        lambda n: ("PCTC" in n.upper() or "OEPP" in n.upper()),
        ncols=13,
    )
    if rows is None:
        return None
    for row in rows:
        label = None
        for v in row:
            if isinstance(v, str) and ADA_RE.search(v) and ADA_9MO_RE.search(v):
                label = v
                break
        if label is None:
            continue
        for v in row:
            if isinstance(v, (int, float)) and 50 < v < 100000:
                return float(v)
    return None


def bucket_unknown_code(code):
    prefix = code[0]
    if prefix == "1":
        return ("Instruction", "Other Instruction", f"Function {code}")
    elif prefix == "2":
        c2 = int(code[:2])
        if c2 == 21:
            return ("Student Support", "Student Support", f"Function {code}")
        elif c2 == 22:
            return ("Instructional Staff Support", "Instructional Staff Support", f"Function {code}")
        elif c2 == 23:
            return ("General Administration", "General Administration", f"Function {code}")
        elif c2 == 24:
            return ("School Administration", "School Administration", f"Function {code}")
        elif c2 == 25:
            return ("Business Services", "Business Services", f"Function {code}")
        elif c2 == 26:
            return ("Central Support", "Central Support", f"Function {code}")
        else:
            return ("Other", "Other", f"Function {code}")
    elif prefix == "3":
        return ("Community Services", "Community Services", f"Function {code}")
    elif prefix == "4":
        return ("Payments to Other Districts", "Payments to Other Districts", f"Function {code}")
    elif prefix == "5":
        return ("Debt Service", "Debt Service", f"Function {code}")
    elif prefix == "6":
        return ("Contingencies", "Contingencies", f"Function {code}")
    return ("Other", "Other", f"Function {code}")


# ---------------------------------------------------------------------------
# Per-district parsing
# ---------------------------------------------------------------------------

def parse_district(folder, name_filter=None):
    """Parse every AFR file in `folder`. Returns:
       year_totals: {fy: {code: total}}, year_ada: {fy: ada}, codes_seen: set
    """
    year_totals = {}
    year_ada = {}
    codes_seen = set()

    for f in sorted(os.listdir(folder)):
        if not f.lower().endswith((".xls", ".xlsx", ".xlsm")):
            continue
        fy = fy_from_filename(f)
        if fy is None:
            continue
        if "Budget" in f:           # FY26 budget arrives as an "AFR26" file
            continue
        if name_filter and name_filter not in f:
            continue

        fp = os.path.join(folder, f)
        records, _ = parse_afr_all(fp, fy)
        if not records:
            continue

        yd = year_totals.setdefault(fy, {})
        for rec in records:
            code = rec["func"]
            codes_seen.add(code)
            yd[code] = yd.get(code, 0.0) + rec["total"]

        ada = extract_ada(fp)
        if ada:
            year_ada[fy] = round(ada, 2)

    return year_totals, year_ada, codes_seen


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    districts = {}        # key -> {"name", "totals": {fy:{code:val}}, "ada": {fy:ada}}
    all_codes_seen = set()
    all_years = set()

    # D65 (top-level files in data/afr/). All D65 AFRs share the RCDTS prefix
    # 05-016-0650-04_ ; this excludes the AFR25_d65.xlsx duplicate and the
    # peers/ subfolder while keeping the year-13 file (named "...No. 65.XLS").
    print("Parsing D65 ...")
    d65_totals, d65_ada, d65_codes = parse_district(AFR_DIR, name_filter="05-016-0650-04_")
    districts[D65_KEY] = {"name": D65_NAME, "totals": d65_totals, "ada": d65_ada}
    all_codes_seen |= d65_codes
    all_years |= set(d65_totals.keys())
    print(f"  {D65_NAME}: years {sorted(d65_totals)}  ADA years {sorted(d65_ada)}")

    # Peers
    for slug, name in PEER_NAMES.items():
        folder = os.path.join(PEER_DIR, slug)
        if not os.path.isdir(folder):
            print(f"!! missing peer folder: {folder}")
            continue
        ptot, pada, pcodes = parse_district(folder)
        districts[slug] = {"name": name, "totals": ptot, "ada": pada}
        all_codes_seen |= pcodes
        all_years |= set(ptot.keys())
        print(f"  {name:28s} years {min(ptot, default='-')}-{max(ptot, default='-')}  "
              f"{len(ptot)} yrs, {len(pada)} ADA yrs")

    years = sorted(all_years)
    year_strs = [str(y) for y in years]
    print(f"\nYears: {years}")

    # Union hierarchy / category metadata
    full_hierarchy = dict(FUNCTION_HIERARCHY)
    for code in all_codes_seen:
        if code not in full_hierarchy and code not in ROLLUP_CODES:
            full_hierarchy[code] = bucket_unknown_code(code)
            print(f"  Unknown code {code} -> {full_hierarchy[code]}")

    categories = {}
    for code in sorted(full_hierarchy.keys()):
        group, subgroup, label = full_hierarchy[code]
        categories[code] = {"label": label, "group": group, "subgroup": subgroup}

    hierarchy = {}
    for code, (group, subgroup, _label) in full_hierarchy.items():
        hierarchy.setdefault(group, {}).setdefault(subgroup, [])
        if code not in hierarchy[group][subgroup]:
            hierarchy[group][subgroup].append(code)
    for group in hierarchy:
        for subgroup in hierarchy[group]:
            hierarchy[group][subgroup].sort()

    # Build per-district numeric payload (arrays aligned to `years`).
    # Only emit codes that have a non-zero value in some year (keeps file small);
    # absent codes read as null in the UI.
    district_order = [D65_KEY] + [s for s in PEER_NAMES if s in districts and s != D65_KEY]
    district_names = {k: districts[k]["name"] for k in district_order}

    values = {}
    totals = {}
    ada = {}
    for key in district_order:
        dt = districts[key]["totals"]
        da = districts[key]["ada"]
        dvals = {}
        for code in full_hierarchy:
            arr = []
            any_nonzero = False
            for y in years:
                if y in dt and code in dt[y]:
                    v = round(dt[y][code], 2)
                    arr.append(v)
                    if v:
                        any_nonzero = True
                else:
                    arr.append(None)
            if any_nonzero:
                dvals[code] = arr
        values[key] = dvals
        totals[key] = [
            round(sum(v for c, v in dt[y].items() if c in full_hierarchy), 2)
            if y in dt else None
            for y in years
        ]
        ada[key] = [da.get(y) for y in years]

    json_data = {
        "years": year_strs,
        "districtOrder": district_order,
        "districtNames": district_names,
        "d65Key": D65_KEY,
        "categories": categories,
        "hierarchy": hierarchy,
        "values": values,
        "totals": totals,
        "ada": ada,
    }

    # --- Validation / summary ---
    print("\n=== Per-district total Ed Fund + ADA (latest common year) ===")
    for key in district_order:
        t = totals[key]
        a = ada[key]
        # latest year that has BOTH a total and an ADA (so per-pupil is shown)
        last = max((i for i in range(len(t)) if t[i] is not None and a[i]), default=None)
        if last is None:
            print(f"  {district_names[key]:28s}  NO DATA")
            continue
        tv = t[last]
        av = a[last]
        pp = (tv / av) if (av and tv) else None
        ada_s = f"{av:,.1f}" if av else "—"
        pp_s = f"${pp:,.0f}" if pp else "—"
        print(f"  {district_names[key]:28s}  FY{year_strs[last]}  "
              f"Ed Fund ${tv:>13,.0f}  ADA {ada_s:>9}  per-pupil {pp_s}")

    html = build_html(json_data)
    out_path = os.path.join(ASSETS_DIR, "budget_explorer.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    size = os.path.getsize(out_path)
    print(f"\nWrote {out_path}  ({size:,} bytes)")


def build_html(json_data):
    data_json = json.dumps(json_data, indent=None, separators=(",", ":"))
    return HTML_TEMPLATE.replace("__BUDGET_DATA__", data_json)


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>D65 Budget Explorer</title>
<script src="https://cdn.plot.ly/plotly-3.3.0.min.js"></script>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
  color: #333; background: #fff; padding: 16px; max-width: 1100px; margin: 0 auto;
}
h1 { font-size: 1.5em; font-weight: 700; margin-bottom: 2px; }
.subtitle { font-size: 0.95em; color: #666; margin-bottom: 14px; }

.controls { margin-bottom: 12px; }
.controls-row { display: flex; flex-wrap: wrap; align-items: center; gap: 6px; margin-bottom: 8px; }
.controls-label { font-size: 0.82em; font-weight: 600; color: #555; margin-right: 4px; white-space: nowrap; }

.preset-btn {
  display: inline-block; padding: 5px 12px; font-size: 0.82em; font-weight: 500;
  border: 1px solid #ccc; border-radius: 4px; background: #f8f8f8; color: #333;
  cursor: pointer; transition: all 0.15s; white-space: nowrap;
}
.preset-btn:hover { background: #e8e8f8; border-color: #99a; }
.preset-btn.active { background: #5e5eb5; color: #fff; border-color: #5e5eb5; }

.toggle-group { display: inline-flex; border: 1px solid #ccc; border-radius: 4px; overflow: hidden; }
.toggle-btn {
  padding: 4px 12px; font-size: 0.82em; font-weight: 500; border: none;
  background: #f8f8f8; color: #555; cursor: pointer; transition: all 0.15s;
  border-right: 1px solid #ccc;
}
.toggle-btn:last-child { border-right: none; }
.toggle-btn:hover { background: #e8e8f8; }
.toggle-btn.active { background: #5e5eb5; color: #fff; }
.toggle-btn:disabled { opacity: 0.4; cursor: not-allowed; background: #f0f0f0; }

#chart { width: 100%; height: 500px; margin-bottom: 8px; }

.mode-banner {
  font-size: 0.82em; padding: 6px 10px; border-radius: 4px; margin-bottom: 12px;
  background: #eef0fb; color: #44488a; border: 1px solid #d4d8f5;
}
.mode-banner b { font-weight: 700; }

/* District selector */
.district-box {
  border: 1px solid #e0e0e0; border-radius: 6px; padding: 10px 12px;
  margin-bottom: 14px; background: #fafafa;
}
.district-grid {
  display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
  gap: 2px 14px;
}
.district-item {
  display: flex; align-items: center; gap: 6px; padding: 2px 0; font-size: 0.84em;
}
.district-item.d65 { font-weight: 700; }
.district-item label { cursor: pointer; }

.tree-header {
  font-size: 0.95em; font-weight: 600; margin-bottom: 8px; color: #333;
  display: flex; align-items: center; gap: 12px;
}
.tree-actions { display: flex; gap: 6px; }
.tree-action-btn {
  font-size: 0.78em; padding: 2px 8px; border: 1px solid #ccc; border-radius: 3px;
  background: #f8f8f8; color: #555; cursor: pointer;
}
.tree-action-btn:hover { background: #e8e8f8; }

.tree-container {
  border: 1px solid #e0e0e0; border-radius: 6px; padding: 10px 14px;
  max-height: 420px; overflow-y: auto; background: #fafafa;
}

.tree-group { margin-bottom: 2px; }
.tree-group-header {
  display: flex; align-items: center; gap: 4px; padding: 3px 0; cursor: pointer;
  font-weight: 600; font-size: 0.88em;
}
.tree-toggle {
  width: 16px; text-align: center; font-size: 0.7em; color: #888;
  cursor: pointer; user-select: none; flex-shrink: 0;
}
.tree-subgroup { margin-left: 20px; margin-bottom: 1px; }
.tree-subgroup-header {
  display: flex; align-items: center; gap: 4px; padding: 2px 0; cursor: pointer;
  font-weight: 500; font-size: 0.85em; color: #444;
}
.tree-leaf {
  display: flex; align-items: center; gap: 4px; padding: 1px 0 1px 20px;
  font-size: 0.83em; color: #555;
}
.tree-leaf .code { color: #999; font-size: 0.9em; }
.tree-children { overflow: hidden; }
.tree-children.collapsed { display: none; }

input[type="checkbox"] { cursor: pointer; accent-color: #5e5eb5; flex-shrink: 0; }

.section-label { font-size: 0.95em; font-weight: 600; margin-bottom: 6px; color: #333; }

.footer {
  margin-top: 16px; padding-top: 10px; border-top: 1px solid #e0e0e0;
  font-size: 0.78em; color: #888; line-height: 1.5;
}

.grouping-row { display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }

@media (max-width: 768px) {
  body { padding: 10px; }
  h1 { font-size: 1.2em; }
  #chart { height: 380px; }
  .tree-container { max-height: 300px; }
}
</style>
</head>
<body>

<h1>D65 Budget Explorer</h1>
<p class="subtitle">Educational Fund Expenditures FY2012–FY2026 (on-behalf/pension payments excluded; D65 FY2026 is adopted budget). Add peer districts to compare.</p>

<div class="controls">
  <div class="controls-row">
    <span class="controls-label">Presets:</span>
    <button class="preset-btn active" data-preset="where-money-goes">Where the Money Goes</button>
    <button class="preset-btn" data-preset="instruction">Instruction Breakdown</button>
    <button class="preset-btn" data-preset="classroom-vs-non">Classroom vs Non-Classroom</button>
    <button class="preset-btn" data-preset="admin">All Administration</button>
    <button class="preset-btn" data-preset="special-ed">Special Ed Total</button>
    <button class="preset-btn" data-preset="student-support">Student Support</button>
    <button class="preset-btn" data-preset="community-payments">Community & Payments</button>
  </div>
  <div class="controls-row">
    <div class="grouping-row">
      <span class="controls-label">Chart:</span>
      <div class="toggle-group" id="chart-type-toggle">
        <button class="toggle-btn active" data-type="area">Area</button>
        <button class="toggle-btn" data-type="line">Line</button>
        <button class="toggle-btn" data-type="bar">Bar</button>
      </div>
    </div>
    <div class="grouping-row">
      <span class="controls-label">Y-Axis:</span>
      <div class="toggle-group" id="metric-toggle">
        <button class="toggle-btn active" data-metric="percent">% of Budget</button>
        <button class="toggle-btn" data-metric="perpupil">Per-Pupil $</button>
        <button class="toggle-btn" data-metric="dollars">Dollars</button>
      </div>
    </div>
  </div>
</div>

<div id="chart"></div>
<div class="mode-banner" id="mode-banner"></div>

<div class="district-box">
  <div class="tree-header" style="margin-bottom:6px;">
    <span>Districts</span>
    <div class="tree-actions">
      <button class="tree-action-btn" id="d65-only-btn">D65 only</button>
      <button class="tree-action-btn" id="all-peers-btn">Select all peers</button>
    </div>
  </div>
  <div class="district-grid" id="district-grid"></div>
</div>

<div class="tree-header">
  <span>Category Selection</span>
  <div class="tree-actions">
    <button class="tree-action-btn" id="select-all-btn">Select All</button>
    <button class="tree-action-btn" id="clear-all-btn">Clear All</button>
  </div>
</div>
<div class="tree-container" id="tree"></div>

<div class="footer">
  <strong>Source:</strong> ISBE Annual Financial Reports (FY2012–FY2025) for each district, and D65's Adopted Budget (FY2026), via districtvitals.com.<br>
  Educational Fund only. On-behalf/pension payments excluded. Category codes follow the ISBE function-code structure.
  <strong>Per-pupil</strong> uses each district's official 9-month ADA from its AFR (the ISBE OEPP denominator).<br>
  <em>Caveat:</em> this is Educational-Fund spending only; districts allocate costs across the Education, Operations &amp; Maintenance, and Transportation funds differently, so cross-district <em>levels</em> are not perfectly apples-to-apples. Comparing the <em>share of budget</em> or trends over time is more robust.
</div>

<script>
const DATA = __BUDGET_DATA__;

const GROUP_ORDER = [
  "Instruction", "Student Support", "Instructional Staff Support",
  "General Administration", "School Administration", "Business Services",
  "Central Support", "Other Support", "Community Services",
  "Payments to Other Districts", "Debt Service", "Contingencies", "Other"
];

const COLORS = [
  "#636EFA", "#EF553B", "#00CC96", "#AB63FA", "#FFA15A",
  "#19D3F3", "#FF6692", "#B6E880", "#FF97FF", "#FECB52",
  "#1F77B4", "#FF7F0E", "#2CA02C", "#D62728", "#9467BD",
  "#8C564B", "#E377C2", "#7F7F7F", "#BCBD22", "#17BECF",
  "#AEC7E8", "#FFBB78", "#98DF8A", "#FF9896", "#C5B0D5"
];

const GROUP_COLORS = {};
GROUP_ORDER.forEach((g, i) => { GROUP_COLORS[g] = COLORS[i % COLORS.length]; });

const D65 = DATA.d65Key;

const state = {
  chartType: "area",
  metric: "percent",
  activePreset: "where-money-goes",
  selectedDistricts: new Set([D65]),
};

// --- helpers ---

function getCodesForGroup(group) {
  const subs = DATA.hierarchy[group] || {};
  const codes = [];
  for (const sg in subs) codes.push(...subs[sg]);
  return codes;
}

function getAllLeafCodes() {
  return Object.keys(DATA.categories);
}

function getCheckedCodes() {
  const codes = [];
  document.querySelectorAll('.leaf-cb:checked').forEach(cb => codes.push(cb.dataset.code));
  return codes;
}

// value of one code for a district at year-index j (null if absent)
function codeVal(dk, code, j) {
  const arr = (DATA.values[dk] || {})[code];
  if (!arr) return null;
  const v = arr[j];
  return (v === null || v === undefined) ? null : v;
}

// sum a set of codes for a district at year-index j; null if no data at all
function sumCodes(dk, codes, j) {
  let sum = 0, has = false;
  for (const c of codes) {
    const v = codeVal(dk, c, j);
    if (v !== null) { sum += v; has = true; }
  }
  return has ? sum : null;
}

// transform a raw dollar value to the active metric for district dk, year j
function applyMetric(dk, dollars, j) {
  if (dollars === null || dollars === undefined) return null;
  if (state.metric === "dollars") return dollars;
  if (state.metric === "percent") {
    const t = (DATA.totals[dk] || [])[j];
    return (t && t > 0) ? (dollars / t * 100) : null;
  }
  if (state.metric === "perpupil") {
    const a = (DATA.ada[dk] || [])[j];
    return (a && a > 0) ? (dollars / a) : null;
  }
  return dollars;
}

function metricAxisTitle() {
  if (state.metric === "percent") return "% of District Ed Fund";
  if (state.metric === "perpupil") return "Per-Pupil Spending ($ / ADA)";
  return "Dollars ($)";
}

function isCompare() {
  return state.selectedDistricts.size >= 2;
}

// --- Presets ---

const PRESETS = {
  "where-money-goes": {
    mode: "grouped",
    groups: function() {
      const g = {};
      GROUP_ORDER.forEach(grp => {
        const codes = getCodesForGroup(grp);
        if (codes.length > 0) g[grp] = codes;
      });
      return g;
    }
  },
  "instruction": {
    mode: "grouped",
    groups: function() {
      const subs = DATA.hierarchy["Instruction"] || {};
      const g = {};
      for (const sg in subs) g[sg] = subs[sg];
      return g;
    }
  },
  "classroom-vs-non": {
    mode: "grouped",
    groups: function() {
      const classroomGroups = ["Instruction", "Instructional Staff Support"];
      const classroom = [], nonClassroom = [];
      GROUP_ORDER.forEach(grp => {
        const codes = getCodesForGroup(grp);
        if (classroomGroups.includes(grp)) classroom.push(...codes);
        else nonClassroom.push(...codes);
      });
      return { "Classroom (Instruction + Instructional Staff)": classroom, "Non-Classroom": nonClassroom };
    }
  },
  "admin": {
    mode: "grouped",
    groups: function() {
      const g = {};
      ["General Administration", "School Administration", "Business Services", "Central Support"].forEach(grp => {
        const subs = DATA.hierarchy[grp] || {};
        for (const sg in subs) g[sg] = subs[sg];
      });
      return g;
    }
  },
  "special-ed": {
    mode: "individual",
    codes: function() { return ["1200", "1225", "1911", "1922", "4120", "4220", "4320"].filter(c => c in DATA.categories); }
  },
  "student-support": {
    mode: "individual",
    codes: function() { return getCodesForGroup("Student Support"); }
  },
  "community-payments": {
    mode: "individual",
    codes: function() {
      return [...getCodesForGroup("Community Services"), ...getCodesForGroup("Payments to Other Districts")];
    }
  },
};

function presetCodes(name) {
  const preset = PRESETS[name];
  if (!preset) return [];
  if (preset.mode === "grouped") {
    const groups = preset.groups();
    const s = new Set();
    for (const g in groups) groups[g].forEach(c => s.add(c));
    return [...s];
  }
  return preset.codes();
}

function applyPreset(name) {
  state.activePreset = name;
  document.querySelectorAll('.preset-btn').forEach(b => b.classList.toggle('active', b.dataset.preset === name));

  const codesInPreset = new Set(presetCodes(name));
  getAllLeafCodes().forEach(code => {
    const cb = document.querySelector(`.leaf-cb[data-code="${code}"]`);
    if (cb) cb.checked = codesInPreset.has(code);
  });
  updateParentCheckboxes();
  updateChart();
}

// --- Trace building ---

// Single-district mode: break the selected district down by category.
function buildSingleTraces(dk) {
  const years = DATA.years;
  const preset = state.activePreset ? PRESETS[state.activePreset] : null;
  const traces = [];

  if (preset && preset.mode === "grouped") {
    const groups = preset.groups();
    // only keep codes the user still has checked
    const checked = new Set(getCheckedCodes());
    Object.keys(groups).forEach((gname, i) => {
      const codes = groups[gname].filter(c => checked.has(c));
      if (!codes.length) return;
      const vals = years.map((_, j) => sumCodes(dk, codes, j));
      if (vals.some(v => v !== null && v > 0))
        traces.push({ name: gname, values: vals, color: COLORS[i % COLORS.length] });
    });
    return traces;
  }

  // individual checked codes
  let colorIdx = 0;
  getCheckedCodes().forEach(code => {
    const cat = DATA.categories[code];
    if (!cat) return;
    const vals = years.map((_, j) => codeVal(dk, code, j));
    if (!vals.some(v => v !== null && v > 0)) return;
    traces.push({
      name: cat.label + " (" + code + ")",
      values: vals,
      color: GROUP_COLORS[cat.group] || COLORS[colorIdx % COLORS.length],
    });
    colorIdx++;
  });
  return traces;
}

// Compare mode: sum the checked categories, one trace per selected district.
function buildCompareTraces() {
  const years = DATA.years;
  const codes = getCheckedCodes();
  const order = DATA.districtOrder.filter(dk => state.selectedDistricts.has(dk));
  const traces = [];
  order.forEach((dk, i) => {
    const vals = years.map((_, j) => sumCodes(dk, codes, j));
    traces.push({
      name: DATA.districtNames[dk] + (dk === D65 ? " ★" : ""),
      values: vals,
      color: dk === D65 ? "#111" : COLORS[i % COLORS.length],
      isD65: dk === D65,
      district: dk,
    });
  });
  return traces;
}

function updateModeBanner() {
  const el = document.getElementById("mode-banner");
  if (isCompare()) {
    const n = state.selectedDistricts.size;
    el.innerHTML = `<b>Compare mode</b> — ${n} districts. Each line is the sum of the checked categories for that district. ` +
      (state.metric === "dollars"
        ? "Showing raw dollars; switch to <b>% of Budget</b> or <b>Per-Pupil $</b> for a fair comparison across district sizes."
        : (state.metric === "percent"
          ? "Showing each category sum as a share of that district's own Educational Fund."
          : "Showing per-pupil spending (category sum ÷ that district's 9-month ADA)."));
  } else {
    const dk = state.selectedDistricts.size ? [...state.selectedDistricts][0] : D65;
    el.innerHTML = `<b>Single-district mode</b> — ${DATA.districtNames[dk]}. ` +
      `Category breakdown. Check additional districts below to compare a category total across districts.`;
  }
}

function updateChart() {
  const years = DATA.years;
  const yearLabels = years.map(y => "FY" + y);
  const compare = isCompare();

  updateModeBanner();

  // In compare mode, stacked area across districts is meaningless -> render as lines.
  let effType = state.chartType;
  if (compare && effType === "area") effType = "line";

  let traces;
  if (compare) {
    traces = buildCompareTraces();
  } else {
    const dk = state.selectedDistricts.size ? [...state.selectedDistricts][0] : D65;
    traces = buildSingleTraces(dk);
  }

  if (traces.length === 0 || traces.every(t => t.values.every(v => v === null))) {
    Plotly.react("chart", [], {
      annotations: [{
        text: compare ? "Select categories to compare across districts"
                      : "Select categories below to display data",
        xref: "paper", yref: "paper", x: 0.5, y: 0.5,
        showarrow: false, font: { size: 16, color: "#999" }
      }],
      xaxis: { visible: false }, yaxis: { visible: false },
      plot_bgcolor: "white", paper_bgcolor: "white", height: 500,
    });
    return;
  }

  const plotTraces = traces.map((t) => {
    const dk = t.district || (state.selectedDistricts.size ? [...state.selectedDistricts][0] : D65);
    const yvals = t.values.map((v, j) => applyMetric(dk, v, j));

    const base = {
      x: yearLabels,
      y: yvals,
      name: t.name,
      marker: { color: t.color },
      line: { color: t.color, width: t.isD65 ? 4 : 2 },
    };

    base.customdata = years.map((yr, j) => {
      const dollars = t.values[j];
      const tot = (DATA.totals[dk] || [])[j];
      const a = (DATA.ada[dk] || [])[j];
      return {
        year: yr,
        dollars: dollars,
        pct: (dollars !== null && tot) ? (dollars / tot * 100) : null,
        pp: (dollars !== null && a) ? (dollars / a) : null,
      };
    });

    const hover = "<b>FY%{customdata.year}</b><br>%{data.name}<br>"
      + "$%{customdata.dollars:,.0f}"
      + " &nbsp;|&nbsp; %{customdata.pct:.1f}% of Ed Fund"
      + " &nbsp;|&nbsp; $%{customdata.pp:,.0f}/pupil<extra></extra>";

    if (effType === "area") {
      base.type = "scatter"; base.mode = "lines"; base.stackgroup = "one";
    } else if (effType === "line") {
      base.type = "scatter"; base.mode = "lines+markers";
    } else {
      base.type = "bar";
    }
    base.hovertemplate = hover;
    base.connectgaps = false;
    return base;
  });

  const layout = {
    plot_bgcolor: "white",
    paper_bgcolor: "white",
    height: 500,
    margin: { l: 75, r: 30, t: 30, b: 80 },
    xaxis: { showgrid: true, gridcolor: "#eee", tickangle: -45 },
    yaxis: {
      showgrid: true, gridcolor: "#eee",
      title: metricAxisTitle(),
      tickformat: state.metric === "percent" ? ".1f" : "$,.0f",
      rangemode: "tozero",
    },
    legend: { orientation: "h", y: -0.25, x: 0.5, xanchor: "center", font: { size: 11 } },
    hovermode: compare ? "closest" : "x unified",
  };

  if (effType === "bar") layout.barmode = compare ? "group" : "stack";

  Plotly.react("chart", plotTraces, layout, { responsive: true });
}

// --- District selector ---

function buildDistrictSelector() {
  const grid = document.getElementById("district-grid");
  grid.innerHTML = "";
  DATA.districtOrder.forEach(dk => {
    const item = document.createElement("div");
    item.className = "district-item" + (dk === D65 ? " d65" : "");
    const cb = document.createElement("input");
    cb.type = "checkbox";
    cb.id = "dist-" + dk;
    cb.dataset.district = dk;
    cb.checked = state.selectedDistricts.has(dk);
    const lbl = document.createElement("label");
    lbl.htmlFor = cb.id;
    lbl.textContent = DATA.districtNames[dk] + (dk === D65 ? " ★" : "");
    cb.addEventListener("change", () => {
      if (cb.checked) state.selectedDistricts.add(dk);
      else state.selectedDistricts.delete(dk);
      if (state.selectedDistricts.size === 0) {
        // never allow an empty selection — fall back to D65
        state.selectedDistricts.add(D65);
        document.getElementById("dist-" + D65).checked = true;
      }
      syncChartToggleForMode();
      updateChart();
    });
    item.appendChild(cb);
    item.appendChild(lbl);
    grid.appendChild(item);
  });
}

function syncChartToggleForMode() {
  // Disable the Area button in compare mode (stacking districts is meaningless).
  const areaBtn = document.querySelector('#chart-type-toggle .toggle-btn[data-type="area"]');
  if (isCompare()) {
    areaBtn.disabled = true;
    if (state.chartType === "area") {
      document.querySelectorAll('#chart-type-toggle .toggle-btn').forEach(b => b.classList.remove('active'));
      document.querySelector('#chart-type-toggle .toggle-btn[data-type="line"]').classList.add('active');
      state.chartType = "line";
    }
  } else {
    areaBtn.disabled = false;
  }
}

// --- Checkbox tree ---

function buildTree() {
  const container = document.getElementById("tree");
  container.innerHTML = "";

  GROUP_ORDER.forEach(group => {
    const subs = DATA.hierarchy[group];
    if (!subs) return;
    const allGroupCodes = getCodesForGroup(group);
    if (allGroupCodes.length === 0) return;

    const groupDiv = document.createElement("div");
    groupDiv.className = "tree-group";

    const header = document.createElement("div");
    header.className = "tree-group-header";
    const toggle = document.createElement("span");
    toggle.className = "tree-toggle"; toggle.textContent = "▼";
    const cb = document.createElement("input");
    cb.type = "checkbox"; cb.className = "group-cb"; cb.dataset.group = group;
    const label = document.createElement("span"); label.textContent = group;
    header.appendChild(toggle); header.appendChild(cb); header.appendChild(label);
    groupDiv.appendChild(header);

    const childrenDiv = document.createElement("div");
    childrenDiv.className = "tree-children";

    const subNames = Object.keys(subs);
    const hasMidLevel = subNames.length > 1 || (subNames.length === 1 && subNames[0] !== group);

    if (hasMidLevel) {
      subNames.forEach(sg => {
        const codes = subs[sg];
        if (codes.length === 1 && sg === codes[0]) {
          appendLeaf(childrenDiv, codes[0]);
        } else {
          const subDiv = document.createElement("div");
          subDiv.className = "tree-subgroup";
          if (codes.length === 1) {
            appendLeaf(subDiv, codes[0]);
          } else {
            const subHeader = document.createElement("div");
            subHeader.className = "tree-subgroup-header";
            const subToggle = document.createElement("span");
            subToggle.className = "tree-toggle"; subToggle.textContent = "▼";
            const subCb = document.createElement("input");
            subCb.type = "checkbox"; subCb.className = "subgroup-cb";
            subCb.dataset.group = group; subCb.dataset.subgroup = sg;
            const subLabel = document.createElement("span"); subLabel.textContent = sg;
            subHeader.appendChild(subToggle); subHeader.appendChild(subCb); subHeader.appendChild(subLabel);
            subDiv.appendChild(subHeader);

            const subChildren = document.createElement("div");
            subChildren.className = "tree-children collapsed";
            codes.forEach(code => appendLeaf(subChildren, code));
            subDiv.appendChild(subChildren);

            subToggle.addEventListener("click", () => {
              subChildren.classList.toggle("collapsed");
              subToggle.textContent = subChildren.classList.contains("collapsed") ? "▶" : "▼";
            });
          }
          childrenDiv.appendChild(subDiv);
        }
      });
    } else {
      const sg = subNames[0];
      subs[sg].forEach(code => appendLeaf(childrenDiv, code));
    }

    groupDiv.appendChild(childrenDiv);
    container.appendChild(groupDiv);

    toggle.addEventListener("click", () => {
      childrenDiv.classList.toggle("collapsed");
      toggle.textContent = childrenDiv.classList.contains("collapsed") ? "▶" : "▼";
    });

    cb.addEventListener("change", () => {
      const checked = cb.checked;
      groupDiv.querySelectorAll(".leaf-cb").forEach(lcb => { lcb.checked = checked; });
      groupDiv.querySelectorAll(".subgroup-cb").forEach(scb => { scb.checked = checked; scb.indeterminate = false; });
      onManualChange();
    });
  });

  document.querySelectorAll(".subgroup-cb").forEach(scb => {
    scb.addEventListener("change", () => {
      const checked = scb.checked;
      const parent = scb.closest(".tree-subgroup");
      parent.querySelectorAll(".leaf-cb").forEach(lcb => { lcb.checked = checked; });
      updateParentCheckboxes();
      onManualChange();
    });
  });

  document.querySelectorAll(".leaf-cb").forEach(lcb => {
    lcb.addEventListener("change", () => { updateParentCheckboxes(); onManualChange(); });
  });
}

function appendLeaf(container, code) {
  const cat = DATA.categories[code];
  if (!cat) return;
  const div = document.createElement("div");
  div.className = "tree-leaf";
  const cb = document.createElement("input");
  cb.type = "checkbox"; cb.className = "leaf-cb"; cb.dataset.code = code;
  const lbl = document.createElement("span"); lbl.textContent = cat.label;
  const codeSpan = document.createElement("span");
  codeSpan.className = "code"; codeSpan.textContent = " (" + code + ")";
  div.appendChild(cb); div.appendChild(lbl); div.appendChild(codeSpan);
  container.appendChild(div);
}

function updateParentCheckboxes() {
  document.querySelectorAll(".subgroup-cb").forEach(scb => {
    const parent = scb.closest(".tree-subgroup");
    const leaves = parent.querySelectorAll(".leaf-cb");
    const checked = parent.querySelectorAll(".leaf-cb:checked");
    scb.checked = checked.length === leaves.length && leaves.length > 0;
    scb.indeterminate = checked.length > 0 && checked.length < leaves.length;
  });
  document.querySelectorAll(".group-cb").forEach(gcb => {
    const group = gcb.closest(".tree-group");
    const leaves = group.querySelectorAll(".leaf-cb");
    const checked = group.querySelectorAll(".leaf-cb:checked");
    gcb.checked = checked.length === leaves.length && leaves.length > 0;
    gcb.indeterminate = checked.length > 0 && checked.length < leaves.length;
  });
}

let debounceTimer = null;
function onManualChange() {
  state.activePreset = null;
  document.querySelectorAll('.preset-btn').forEach(b => b.classList.remove('active'));
  if (debounceTimer) clearTimeout(debounceTimer);
  debounceTimer = setTimeout(updateChart, 100);
}

// --- Toggle handlers ---

document.querySelectorAll('#chart-type-toggle .toggle-btn').forEach(btn => {
  btn.addEventListener("click", () => {
    if (btn.disabled) return;
    document.querySelectorAll('#chart-type-toggle .toggle-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    state.chartType = btn.dataset.type;
    updateChart();
  });
});

document.querySelectorAll('#metric-toggle .toggle-btn').forEach(btn => {
  btn.addEventListener("click", () => {
    document.querySelectorAll('#metric-toggle .toggle-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    state.metric = btn.dataset.metric;
    updateChart();
  });
});

document.querySelectorAll('.preset-btn').forEach(btn => {
  btn.addEventListener("click", () => applyPreset(btn.dataset.preset));
});

document.getElementById("select-all-btn").addEventListener("click", () => {
  document.querySelectorAll(".leaf-cb").forEach(cb => { cb.checked = true; });
  updateParentCheckboxes(); onManualChange();
});
document.getElementById("clear-all-btn").addEventListener("click", () => {
  document.querySelectorAll(".leaf-cb").forEach(cb => { cb.checked = false; });
  updateParentCheckboxes(); onManualChange();
});

document.getElementById("d65-only-btn").addEventListener("click", () => {
  state.selectedDistricts = new Set([D65]);
  document.querySelectorAll('.district-item input').forEach(cb => { cb.checked = (cb.dataset.district === D65); });
  syncChartToggleForMode(); updateChart();
});
document.getElementById("all-peers-btn").addEventListener("click", () => {
  DATA.districtOrder.forEach(dk => state.selectedDistricts.add(dk));
  document.querySelectorAll('.district-item input').forEach(cb => { cb.checked = true; });
  syncChartToggleForMode(); updateChart();
});

// --- Init ---
buildDistrictSelector();
buildTree();
applyPreset("where-money-goes");
</script>
</body>
</html>
"""


if __name__ == "__main__":
    main()
